"""Standalone dashboard for MetLife financial ratios using Tkinter and ttkbootstrap."""

from __future__ import annotations

import argparse
import re
import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import ttk as tk_ttk

import pandas as pd
import ttkbootstrap as ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


FONT_HEADER = ("Segoe UI Semibold", 20)
FONT_TITLE = ("Segoe UI Semibold", 13)
FONT_BODY = ("Segoe UI", 10)
FONT_SMALL = ("Segoe UI", 9)
COLOR_POSITIVE = "#0d6efd"
COLOR_NEGATIVE = "#dc3545"
COLOR_NEUTRAL = "#6c757d"
COLOR_ACCENT = "#198754"
CATEGORY_ORDER = [
    "Liquidez",
    "Solvencia y Apalancamiento",
    "Suficiencia de la Prima",
    "Reaseguro",
    "Rentabilidad",
]


@dataclass
class DashboardData:
    workbook_path: Path
    ratios_wide: pd.DataFrame
    ratios_long: pd.DataFrame
    accounts_wide: pd.DataFrame
    accounts_long: pd.DataFrame
    years: list[int]
    categories: list[str]


class ExcelDashboardLoader:
    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"No se encontro el workbook: {self.workbook_path}")
        self.table_refs = self._resolve_named_tables()

    def load(self) -> DashboardData:
        ratios_wide = self._read_named_table("Razones_anuales")
        accounts_wide = self._read_named_table("Cuentas_anuales")

        ratios_wide = self._clean_ratios_table(ratios_wide)
        accounts_wide = self._clean_accounts_table(accounts_wide)

        years = self._detect_years(ratios_wide)
        year_columns = [str(year) for year in years]

        ratios_long = ratios_wide.melt(
            id_vars=["Categoria", "Razon"],
            value_vars=year_columns,
            var_name="Anio",
            value_name="Valor",
        )
        ratios_long["Anio"] = ratios_long["Anio"].astype(int)
        ratios_long["Valor"] = pd.to_numeric(ratios_long["Valor"], errors="coerce")
        ratios_long = ratios_long.dropna(subset=["Valor"]).reset_index(drop=True)

        accounts_long = accounts_wide.melt(
            id_vars=["Cuentas"],
            value_vars=year_columns,
            var_name="Anio",
            value_name="Valor",
        )
        accounts_long["Anio"] = accounts_long["Anio"].astype(int)
        accounts_long["Valor"] = pd.to_numeric(accounts_long["Valor"], errors="coerce")
        accounts_long = accounts_long.dropna(subset=["Valor"]).reset_index(drop=True)

        workbook_categories = ratios_wide["Categoria"].dropna().astype(str).tolist()
        unique_categories = list(dict.fromkeys(workbook_categories))
        ordered_categories = [category for category in CATEGORY_ORDER if category in unique_categories]
        ordered_categories.extend([category for category in unique_categories if category not in ordered_categories])

        return DashboardData(
            workbook_path=self.workbook_path,
            ratios_wide=ratios_wide,
            ratios_long=ratios_long,
            accounts_wide=accounts_wide,
            accounts_long=accounts_long,
            years=years,
            categories=ordered_categories,
        )

    def _resolve_named_tables(self) -> dict[str, tuple[str, str]]:
        workbook = load_workbook(self.workbook_path, data_only=True)
        table_refs: dict[str, tuple[str, str]] = {}
        try:
            for sheet in workbook.worksheets:
                for table_name in sheet.tables.keys():
                    table_refs[table_name] = (sheet.title, sheet.tables[table_name].ref)
        finally:
            workbook.close()
        return table_refs

    def _read_named_table(self, table_name: str) -> pd.DataFrame:
        if table_name not in self.table_refs:
            available = ", ".join(sorted(self.table_refs.keys()))
            raise ValueError(f"No existe la tabla '{table_name}'. Tablas detectadas: {available}")

        sheet_name, table_ref = self.table_refs[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table_ref)
        usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"
        dataframe = pd.read_excel(
            self.workbook_path,
            sheet_name=sheet_name,
            header=min_row - 1,
            nrows=max_row - min_row,
            usecols=usecols,
        )
        dataframe = dataframe.dropna(how="all").dropna(axis=1, how="all")
        dataframe.columns = [self._normalize_column_name(column) for column in dataframe.columns]
        return dataframe

    @staticmethod
    def _normalize_column_name(column: object) -> str:
        text = str(column).strip()
        match = re.fullmatch(r"(\d{4})(?:\.\d+)?", text)
        if match:
            return match.group(1)
        return text

    @staticmethod
    def _detect_years(dataframe: pd.DataFrame) -> list[int]:
        years = sorted(int(column) for column in dataframe.columns if re.fullmatch(r"\d{4}", str(column)))
        if not years:
            raise ValueError("No se detectaron columnas de anos en la tabla de razones.")
        return years

    def _clean_ratios_table(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        required_columns = {"Categoria", "Razon"}
        if not required_columns.issubset(dataframe.columns):
            raise ValueError(f"La tabla de razones debe contener las columnas {sorted(required_columns)}")

        cleaned = dataframe.copy()
        cleaned["Categoria"] = cleaned["Categoria"].astype(str).str.strip()
        cleaned["Razon"] = cleaned["Razon"].astype(str).str.strip()
        cleaned = cleaned.loc[(cleaned["Categoria"] != "") & (cleaned["Razon"] != "")].reset_index(drop=True)

        for year in self._detect_years(cleaned):
            cleaned[str(year)] = pd.to_numeric(cleaned[str(year)], errors="coerce")

        return cleaned

    def _clean_accounts_table(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        if "Cuentas" not in dataframe.columns:
            raise ValueError("La tabla de cuentas debe contener la columna 'Cuentas'.")

        cleaned = dataframe.copy()
        cleaned["Cuentas"] = cleaned["Cuentas"].astype(str).str.strip()
        cleaned = cleaned.loc[cleaned["Cuentas"] != ""].reset_index(drop=True)

        year_columns = [column for column in cleaned.columns if re.fullmatch(r"\d{4}", str(column))]
        if not year_columns:
            raise ValueError("La tabla de cuentas no contiene columnas anuales validas.")

        for year_column in year_columns:
            cleaned[year_column] = pd.to_numeric(cleaned[year_column], errors="coerce")

        return cleaned


def format_ratio(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    return f"{float(value):,.2f}"


def format_delta(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    return f"{float(value):+,.2f}"


def format_compact_number(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "N/D"

    absolute = abs(float(value))
    if absolute >= 1_000_000_000_000:
        return f"{float(value) / 1_000_000_000_000:,.2f}T"
    if absolute >= 1_000_000_000:
        return f"{float(value) / 1_000_000_000:,.2f}B"
    if absolute >= 1_000_000:
        return f"{float(value) / 1_000_000:,.2f}M"
    if absolute >= 1_000:
        return f"{float(value) / 1_000:,.2f}K"
    return f"{float(value):,.2f}"


def short_label(text: str, limit: int = 42) -> str:
    clean_text = str(text).strip()
    if len(clean_text) <= limit:
        return clean_text
    return clean_text[: limit - 3].rstrip() + "..."


class MetLifeDashboardApp:
    def __init__(self, data: DashboardData):
        self.data = data
        self.latest_year = max(self.data.years)
        self.base_year = min(self.data.years)
        self.root = ttk.Window(themename="cosmo")
        self.root.title("Dashboard Ejecutivo - MetLife")
        self.root.geometry("1680x980")
        self.root.minsize(1400, 860)
        self.category_state: dict[str, dict[str, object]] = {}
        self.accounts_state: dict[str, object] = {}
        self._build_ui()

    def run(self) -> None:
        self.root.mainloop()

    def _build_ui(self) -> None:
        outer = ttk.Frame(self.root, padding=18)
        outer.pack(fill="both", expand=True)

        header = ttk.Frame(outer)
        header.pack(fill="x")
        ttk.Label(header, text="Dashboard Ejecutivo de Razones Financieras", font=FONT_HEADER).pack(anchor="w")
        ttk.Label(
            header,
            text="Empresa: MetLife | Periodo: 2021-2025 | Fuente: PIA_main_razones.xlsx",
            font=FONT_BODY,
        ).pack(anchor="w", pady=(4, 0))
        ttk.Label(
            header,
            text="Analisis historico de liquidez, solvencia y apalancamiento, suficiencia de la prima, reaseguro y rentabilidad.",
            font=FONT_SMALL,
        ).pack(anchor="w", pady=(4, 10))

        notebook = ttk.Notebook(outer)
        notebook.pack(fill="both", expand=True)

        self._build_overview_tab(notebook)
        for category in self.data.categories:
            self._build_category_tab(notebook, category)
        self._build_accounts_tab(notebook)

    @staticmethod
    def _clear_frame(frame: tk.Misc) -> None:
        for child in frame.winfo_children():
            child.destroy()

    def _draw_figure(self, frame: tk.Misc, figure: Figure) -> None:
        self._clear_frame(frame)
        canvas = FigureCanvasTkAgg(figure, master=frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        frame.canvas = canvas

    def _populate_cards(self, frame: tk.Misc, cards: list[dict[str, str]]) -> None:
        self._clear_frame(frame)
        for column, card in enumerate(cards):
            frame.columnconfigure(column, weight=1)
            card_frame = tk.Frame(frame, bd=1, relief="solid", padx=12, pady=10, bg="white")
            card_frame.grid(row=0, column=column, sticky="nsew", padx=6, pady=6)

            tk.Label(
                card_frame,
                text=card["title"],
                font=("Segoe UI", 10, "bold"),
                bg="white",
                anchor="w",
                justify="left",
                wraplength=210,
            ).pack(anchor="w")
            tk.Label(
                card_frame,
                text=card["value"],
                font=("Segoe UI Semibold", 18),
                bg="white",
                fg=card.get("color", "#212529"),
            ).pack(anchor="w", pady=(8, 0))

            detail_text = card.get("detail")
            if detail_text:
                tk.Label(
                    card_frame,
                    text=detail_text,
                    font=FONT_SMALL,
                    bg="white",
                    fg="#495057",
                    anchor="w",
                    justify="left",
                    wraplength=230,
                ).pack(anchor="w", pady=(6, 0))

    def _draw_treeview(
        self,
        frame: tk.Misc,
        dataframe: pd.DataFrame,
        numeric_columns: set[str] | None = None,
        compact_numeric_columns: set[str] | None = None,
        height: int = 12,
    ) -> None:
        self._clear_frame(frame)
        numeric_columns = numeric_columns or set()
        compact_numeric_columns = compact_numeric_columns or set()

        tree_frame = ttk.Frame(frame)
        tree_frame.pack(fill="both", expand=True)

        columns = list(dataframe.columns)
        tree = tk_ttk.Treeview(tree_frame, columns=columns, show="headings", height=height)
        tree.grid(row=0, column=0, sticky="nsew")

        y_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
        x_scroll.grid(row=1, column=0, sticky="ew")

        tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        for column in columns:
            width = 160 if column not in numeric_columns and column not in compact_numeric_columns else 95
            anchor = "center" if column in numeric_columns or column in compact_numeric_columns else "w"
            tree.heading(column, text=column)
            tree.column(column, width=width, minwidth=80, anchor=anchor, stretch=True)

        for row in dataframe.itertuples(index=False):
            values = []
            for column, value in zip(columns, row):
                if column in compact_numeric_columns:
                    values.append(format_compact_number(value))
                elif column in numeric_columns:
                    values.append(format_ratio(value))
                else:
                    values.append(value)
            tree.insert("", "end", values=values)

    def _build_overview_tab(self, notebook: ttk.Notebook) -> None:
        tab = ttk.Frame(notebook, padding=16)
        notebook.add(tab, text="Portada Ejecutiva")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(2, weight=1)
        tab.rowconfigure(3, weight=1)

        intro = ttk.Frame(tab)
        intro.grid(row=0, column=0, sticky="ew")
        ttk.Label(intro, text="MetLife - Vision general del periodo 2021-2025", font=FONT_TITLE).pack(anchor="w")
        ttk.Label(
            intro,
            text="La portada concentra el panorama global de las razones financieras y los principales movimientos historicos del periodo.",
            font=FONT_BODY,
        ).pack(anchor="w", pady=(4, 0))

        cards_frame = ttk.Frame(tab)
        cards_frame.grid(row=1, column=0, sticky="ew", pady=(12, 10))

        delta_snapshot = self.data.ratios_wide[["Categoria", "Razon", str(self.base_year), str(self.latest_year)]].copy()
        delta_snapshot["Delta_2025_2021"] = delta_snapshot[str(self.latest_year)] - delta_snapshot[str(self.base_year)]
        max_delta_row = delta_snapshot.loc[delta_snapshot["Delta_2025_2021"].idxmax()]
        min_delta_row = delta_snapshot.loc[delta_snapshot["Delta_2025_2021"].idxmin()]

        overview_cards = [
            {
                "title": "Categorias analizadas",
                "value": str(len(self.data.categories)),
                "detail": "Liquidez, solvencia y apalancamiento, suficiencia de la prima, reaseguro y rentabilidad.",
                "color": COLOR_POSITIVE,
            },
            {
                "title": "Razones disponibles",
                "value": str(len(self.data.ratios_wide)),
                "detail": "Indicadores historicos ya calculados para los cinco anos del analisis.",
                "color": COLOR_ACCENT,
            },
            {
                "title": "Cuentas base",
                "value": str(len(self.data.accounts_wide)),
                "detail": "Rubros auxiliares disponibles para respaldo del calculo de razones.",
                "color": COLOR_NEUTRAL,
            },
            {
                "title": "Ultimo ano disponible",
                "value": str(self.latest_year),
                "detail": "La comparacion ejecutiva usa 2021 como linea base y 2025 como referencia mas reciente.",
                "color": COLOR_POSITIVE,
            },
            {
                "title": "Mayor avance 2021-2025",
                "value": short_label(max_delta_row["Razon"], 28),
                "detail": f"Cambio: {format_delta(max_delta_row['Delta_2025_2021'])}",
                "color": COLOR_POSITIVE,
            },
            {
                "title": "Mayor caida 2021-2025",
                "value": short_label(min_delta_row["Razon"], 28),
                "detail": f"Cambio: {format_delta(min_delta_row['Delta_2025_2021'])}",
                "color": COLOR_NEGATIVE,
            },
        ]
        self._populate_cards(cards_frame, overview_cards)

        chart_row = ttk.Frame(tab)
        chart_row.grid(row=2, column=0, sticky="nsew", pady=(8, 8))
        chart_row.columnconfigure(0, weight=1)
        chart_row.columnconfigure(1, weight=1)
        chart_row.rowconfigure(0, weight=1)

        trend_section = ttk.Labelframe(chart_row, text="Promedio simple por categoria", padding=10)
        trend_section.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        delta_section = ttk.Labelframe(chart_row, text="Top cambios 2021-2025", padding=10)
        delta_section.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        self._draw_figure(trend_section, self._make_category_average_figure())
        self._draw_figure(delta_section, self._make_overview_delta_figure())

        detail_section = ttk.Labelframe(tab, text="Detalle anual de razones", padding=10)
        detail_section.grid(row=3, column=0, sticky="nsew")
        detail_snapshot = delta_snapshot[["Categoria", "Razon", str(self.base_year), str(self.latest_year), "Delta_2025_2021"]].copy()
        detail_snapshot = detail_snapshot.rename(columns={"Delta_2025_2021": f"Delta {self.latest_year}-{self.base_year}"})
        full_snapshot = self.data.ratios_wide.copy()
        full_snapshot = full_snapshot.merge(detail_snapshot[["Razon", f"Delta {self.latest_year}-{self.base_year}"]], on="Razon", how="left")
        self._draw_treeview(
            detail_section,
            full_snapshot[["Categoria", "Razon", *[str(year) for year in self.data.years], f"Delta {self.latest_year}-{self.base_year}"]],
            numeric_columns=set([str(year) for year in self.data.years] + [f"Delta {self.latest_year}-{self.base_year}"]),
            height=13,
        )

    def _build_category_tab(self, notebook: ttk.Notebook, category: str) -> None:
        tab = ttk.Frame(notebook, padding=16)
        notebook.add(tab, text=category)
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)
        tab.rowconfigure(4, weight=1)

        category_wide = self.data.ratios_wide.loc[self.data.ratios_wide["Categoria"] == category].copy()
        ratio_list = category_wide["Razon"].tolist()

        title_frame = ttk.Frame(tab)
        title_frame.grid(row=0, column=0, sticky="ew")
        ttk.Label(title_frame, text=f"{category} - MetLife", font=FONT_TITLE).pack(anchor="w")
        ttk.Label(
            title_frame,
            text="Comparativos historicos, tarjetas KPI y detalle anual para la categoria seleccionada.",
            font=FONT_BODY,
        ).pack(anchor="w", pady=(4, 0))

        controls = ttk.Frame(tab)
        controls.grid(row=1, column=0, sticky="ew", pady=(12, 10))

        year_var = tk.StringVar(value=str(self.latest_year))
        ratio_var = tk.StringVar(value=ratio_list[0])
        ttk.Label(controls, text="Ano de referencia", font=FONT_BODY).pack(side="left")
        year_combo = ttk.Combobox(controls, values=[str(year) for year in self.data.years], textvariable=year_var, width=10, state="readonly")
        year_combo.pack(side="left", padx=(8, 18))
        ttk.Label(controls, text="Razon destacada", font=FONT_BODY).pack(side="left")
        ratio_combo = ttk.Combobox(controls, values=ratio_list, textvariable=ratio_var, width=48, state="readonly")
        ratio_combo.pack(side="left", padx=(8, 0), fill="x", expand=True)

        cards_frame = ttk.Frame(tab)
        cards_frame.grid(row=2, column=0, sticky="ew")

        chart_row = ttk.Frame(tab)
        chart_row.grid(row=3, column=0, sticky="nsew", pady=(8, 8))
        chart_row.columnconfigure(0, weight=1)
        chart_row.columnconfigure(1, weight=1)
        chart_row.rowconfigure(0, weight=1)

        trend_section = ttk.Labelframe(chart_row, text="Tendencia de la razon seleccionada", padding=10)
        trend_section.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ranking_section = ttk.Labelframe(chart_row, text="Comparativo de la categoria por ano", padding=10)
        ranking_section.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        detail_row = ttk.Frame(tab)
        detail_row.grid(row=4, column=0, sticky="nsew")
        detail_row.columnconfigure(0, weight=1)
        detail_row.columnconfigure(1, weight=1)
        detail_row.rowconfigure(0, weight=1)

        delta_section = ttk.Labelframe(detail_row, text="Cambio 2021-2025 dentro de la categoria", padding=10)
        delta_section.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        table_section = ttk.Labelframe(detail_row, text="Detalle anual", padding=10)
        table_section.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        self.category_state[category] = {
            "wide": category_wide,
            "long": self.data.ratios_long.loc[self.data.ratios_long["Categoria"] == category].copy(),
            "year_var": year_var,
            "ratio_var": ratio_var,
            "cards_frame": cards_frame,
            "trend_section": trend_section,
            "ranking_section": ranking_section,
            "delta_section": delta_section,
            "table_section": table_section,
        }

        year_combo.bind("<<ComboboxSelected>>", lambda _event, selected_category=category: self._update_category_tab(selected_category))
        ratio_combo.bind("<<ComboboxSelected>>", lambda _event, selected_category=category: self._update_category_tab(selected_category))

        self._update_category_tab(category)

    def _build_accounts_tab(self, notebook: ttk.Notebook) -> None:
        tab = ttk.Frame(notebook, padding=16)
        notebook.add(tab, text="Cuentas Base")
        tab.columnconfigure(0, weight=1)
        tab.rowconfigure(3, weight=1)
        tab.rowconfigure(4, weight=1)

        title_frame = ttk.Frame(tab)
        title_frame.grid(row=0, column=0, sticky="ew")
        ttk.Label(title_frame, text="Cuentas Base - MetLife", font=FONT_TITLE).pack(anchor="w")
        ttk.Label(
            title_frame,
            text="Vista de soporte para revisar la evolucion historica de las cuentas utilizadas en el analisis.",
            font=FONT_BODY,
        ).pack(anchor="w", pady=(4, 0))

        controls = ttk.Frame(tab)
        controls.grid(row=1, column=0, sticky="ew", pady=(12, 10))

        search_var = tk.StringVar()
        year_var = tk.StringVar(value=str(self.latest_year))
        account_var = tk.StringVar(value=self.data.accounts_wide["Cuentas"].iloc[0])

        ttk.Label(controls, text="Buscar cuenta", font=FONT_BODY).pack(side="left")
        search_entry = ttk.Entry(controls, textvariable=search_var, width=28)
        search_entry.pack(side="left", padx=(8, 18))
        ttk.Label(controls, text="Ano de referencia", font=FONT_BODY).pack(side="left")
        year_combo = ttk.Combobox(controls, values=[str(year) for year in self.data.years], textvariable=year_var, width=10, state="readonly")
        year_combo.pack(side="left", padx=(8, 18))
        ttk.Label(controls, text="Cuenta destacada", font=FONT_BODY).pack(side="left")
        account_combo = ttk.Combobox(
            controls,
            values=self.data.accounts_wide["Cuentas"].tolist(),
            textvariable=account_var,
            width=42,
            state="readonly",
        )
        account_combo.pack(side="left", padx=(8, 0), fill="x", expand=True)

        cards_frame = ttk.Frame(tab)
        cards_frame.grid(row=2, column=0, sticky="ew")

        chart_row = ttk.Frame(tab)
        chart_row.grid(row=3, column=0, sticky="nsew", pady=(8, 8))
        chart_row.columnconfigure(0, weight=1)
        chart_row.columnconfigure(1, weight=1)
        chart_row.rowconfigure(0, weight=1)

        trend_section = ttk.Labelframe(chart_row, text="Tendencia de la cuenta seleccionada", padding=10)
        trend_section.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        ranking_section = ttk.Labelframe(chart_row, text="Top cuentas por ano", padding=10)
        ranking_section.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        table_section = ttk.Labelframe(tab, text="Detalle anual de cuentas", padding=10)
        table_section.grid(row=4, column=0, sticky="nsew")

        self.accounts_state = {
            "search_var": search_var,
            "year_var": year_var,
            "account_var": account_var,
            "account_combo": account_combo,
            "cards_frame": cards_frame,
            "trend_section": trend_section,
            "ranking_section": ranking_section,
            "table_section": table_section,
        }

        search_var.trace_add("write", lambda *_args: self._sync_account_options())
        year_combo.bind("<<ComboboxSelected>>", lambda _event: self._update_accounts_tab())
        account_combo.bind("<<ComboboxSelected>>", lambda _event: self._update_accounts_tab())

        self._sync_account_options()

    def _make_category_average_figure(self) -> Figure:
        figure = Figure(figsize=(7.8, 4.2), dpi=100)
        axis = figure.add_subplot(111)

        grouped = self.data.ratios_long.groupby(["Categoria", "Anio"], as_index=False)["Valor"].mean()
        for category in self.data.categories:
            category_frame = grouped.loc[grouped["Categoria"] == category]
            axis.plot(category_frame["Anio"], category_frame["Valor"], marker="o", linewidth=2, label=category)

        axis.set_title("Promedio simple por categoria y ano")
        axis.set_xlabel("Ano")
        axis.set_ylabel("Valor promedio")
        axis.grid(axis="y", alpha=0.3)
        axis.legend(loc="best", fontsize=8)
        figure.tight_layout()
        return figure

    def _make_overview_delta_figure(self) -> Figure:
        figure = Figure(figsize=(7.8, 4.2), dpi=100)
        axis = figure.add_subplot(111)

        delta_frame = self.data.ratios_wide[["Razon", str(self.base_year), str(self.latest_year)]].copy()
        delta_frame["Delta"] = delta_frame[str(self.latest_year)] - delta_frame[str(self.base_year)]
        delta_frame = delta_frame.reindex(delta_frame["Delta"].abs().sort_values(ascending=False).index).head(10)
        delta_frame = delta_frame.sort_values("Delta")

        colors = [COLOR_NEGATIVE if value < 0 else COLOR_POSITIVE for value in delta_frame["Delta"]]
        axis.barh([short_label(label, 34) for label in delta_frame["Razon"]], delta_frame["Delta"], color=colors)
        axis.axvline(0, color=COLOR_NEUTRAL, linewidth=1)
        axis.set_title("Variacion absoluta entre 2021 y 2025")
        axis.set_xlabel("Cambio absoluto")
        axis.grid(axis="x", alpha=0.3)
        figure.tight_layout()
        return figure

    def _update_category_tab(self, category: str) -> None:
        state = self.category_state[category]
        wide = state["wide"]
        long_frame = state["long"]
        selected_year = int(state["year_var"].get())
        selected_ratio = state["ratio_var"].get()

        ratio_history = long_frame.loc[long_frame["Razon"] == selected_ratio].sort_values("Anio")
        selected_row = ratio_history.loc[ratio_history["Anio"] == selected_year]
        selected_value = selected_row["Valor"].iloc[0] if not selected_row.empty else None
        base_row = ratio_history.loc[ratio_history["Anio"] == self.base_year]
        base_value = base_row["Valor"].iloc[0] if not base_row.empty else None
        delta_value = selected_value - base_value if selected_value is not None and base_value is not None else None
        mean_value = ratio_history["Valor"].mean() if not ratio_history.empty else None
        max_row = ratio_history.loc[ratio_history["Valor"].idxmax()] if not ratio_history.empty else None
        min_row = ratio_history.loc[ratio_history["Valor"].idxmin()] if not ratio_history.empty else None

        cards = [
            {
                "title": f"Valor en {selected_year}",
                "value": format_ratio(selected_value),
                "detail": short_label(selected_ratio, 60),
                "color": COLOR_POSITIVE,
            },
            {
                "title": f"Cambio vs {self.base_year}",
                "value": format_delta(delta_value),
                "detail": "Variacion absoluta de la razon seleccionada.",
                "color": COLOR_NEGATIVE if delta_value is not None and delta_value < 0 else COLOR_POSITIVE,
            },
            {
                "title": "Promedio 5 anos",
                "value": format_ratio(mean_value),
                "detail": "Promedio simple de la razon en todo el horizonte historico.",
                "color": COLOR_ACCENT,
            },
            {
                "title": "Maximo historico",
                "value": format_ratio(max_row["Valor"] if max_row is not None else None),
                "detail": f"Ano: {int(max_row['Anio'])}" if max_row is not None else "N/D",
                "color": COLOR_POSITIVE,
            },
            {
                "title": "Minimo historico",
                "value": format_ratio(min_row["Valor"] if min_row is not None else None),
                "detail": f"Ano: {int(min_row['Anio'])}" if min_row is not None else "N/D",
                "color": COLOR_NEGATIVE,
            },
        ]
        self._populate_cards(state["cards_frame"], cards)

        self._draw_figure(state["trend_section"], self._make_selected_ratio_trend_figure(ratio_history, selected_ratio))
        self._draw_figure(state["ranking_section"], self._make_category_year_ranking_figure(long_frame, selected_year))
        self._draw_figure(state["delta_section"], self._make_category_delta_figure(wide))

        detail_frame = wide[["Razon", *[str(year) for year in self.data.years]]].copy()
        detail_frame[f"Delta {self.latest_year}-{self.base_year}"] = detail_frame[str(self.latest_year)] - detail_frame[str(self.base_year)]
        self._draw_treeview(
            state["table_section"],
            detail_frame,
            numeric_columns=set([str(year) for year in self.data.years] + [f"Delta {self.latest_year}-{self.base_year}"]),
            height=10,
        )

    def _make_selected_ratio_trend_figure(self, ratio_history: pd.DataFrame, selected_ratio: str) -> Figure:
        figure = Figure(figsize=(7.8, 4.2), dpi=100)
        axis = figure.add_subplot(111)

        axis.plot(ratio_history["Anio"], ratio_history["Valor"], marker="o", linewidth=2.5, color=COLOR_POSITIVE)
        for year, value in zip(ratio_history["Anio"], ratio_history["Valor"]):
            axis.annotate(format_ratio(value), (year, value), textcoords="offset points", xytext=(0, 8), ha="center", fontsize=8)

        axis.set_title(short_label(selected_ratio, 60))
        axis.set_xlabel("Ano")
        axis.set_ylabel("Valor")
        axis.grid(axis="y", alpha=0.3)
        figure.tight_layout()
        return figure

    def _make_category_year_ranking_figure(self, long_frame: pd.DataFrame, selected_year: int) -> Figure:
        figure = Figure(figsize=(7.8, 4.2), dpi=100)
        axis = figure.add_subplot(111)

        year_frame = long_frame.loc[long_frame["Anio"] == selected_year].sort_values("Valor")
        axis.barh([short_label(label, 36) for label in year_frame["Razon"]], year_frame["Valor"], color=COLOR_ACCENT)
        axis.set_title(f"Comparativo de razones en {selected_year}")
        axis.set_xlabel("Valor")
        axis.grid(axis="x", alpha=0.3)
        figure.tight_layout()
        return figure

    def _make_category_delta_figure(self, wide: pd.DataFrame) -> Figure:
        figure = Figure(figsize=(7.8, 4.2), dpi=100)
        axis = figure.add_subplot(111)

        delta_frame = wide[["Razon", str(self.base_year), str(self.latest_year)]].copy()
        delta_frame["Delta"] = delta_frame[str(self.latest_year)] - delta_frame[str(self.base_year)]
        delta_frame = delta_frame.sort_values("Delta")
        colors = [COLOR_NEGATIVE if value < 0 else COLOR_POSITIVE for value in delta_frame["Delta"]]

        axis.barh([short_label(label, 36) for label in delta_frame["Razon"]], delta_frame["Delta"], color=colors)
        axis.axvline(0, color=COLOR_NEUTRAL, linewidth=1)
        axis.set_title(f"Cambio {self.base_year}-{self.latest_year}")
        axis.set_xlabel("Delta")
        axis.grid(axis="x", alpha=0.3)
        figure.tight_layout()
        return figure

    def _sync_account_options(self) -> None:
        state = self.accounts_state
        query = state["search_var"].get().strip().lower()
        available_accounts = self.data.accounts_wide["Cuentas"].tolist()
        filtered_accounts = [account for account in available_accounts if query in account.lower()]

        state["account_combo"]["values"] = filtered_accounts
        current_account = state["account_var"].get()
        if filtered_accounts and current_account not in filtered_accounts:
            state["account_var"].set(filtered_accounts[0])
        elif not filtered_accounts:
            state["account_var"].set("")

        self._update_accounts_tab()

    def _update_accounts_tab(self) -> None:
        state = self.accounts_state
        selected_year = int(state["year_var"].get())
        selected_account = state["account_var"].get()
        query = state["search_var"].get().strip().lower()

        filtered_wide = self.data.accounts_wide.copy()
        if query:
            filtered_wide = filtered_wide.loc[filtered_wide["Cuentas"].str.lower().str.contains(query, regex=False)]

        if filtered_wide.empty:
            cards = [
                {"title": "Estado", "value": "Sin datos", "detail": "No hay cuentas que coincidan con la busqueda.", "color": COLOR_NEGATIVE},
            ]
            self._populate_cards(state["cards_frame"], cards)
            self._clear_frame(state["trend_section"])
            self._clear_frame(state["ranking_section"])
            self._clear_frame(state["table_section"])
            return

        if not selected_account or selected_account not in filtered_wide["Cuentas"].tolist():
            selected_account = filtered_wide["Cuentas"].iloc[0]
            state["account_var"].set(selected_account)

        filtered_long = self.data.accounts_long.loc[self.data.accounts_long["Cuentas"].isin(filtered_wide["Cuentas"])]
        account_history = filtered_long.loc[filtered_long["Cuentas"] == selected_account].sort_values("Anio")
        selected_row = account_history.loc[account_history["Anio"] == selected_year]
        selected_value = selected_row["Valor"].iloc[0] if not selected_row.empty else None
        base_row = account_history.loc[account_history["Anio"] == self.base_year]
        base_value = base_row["Valor"].iloc[0] if not base_row.empty else None
        delta_value = selected_value - base_value if selected_value is not None and base_value is not None else None
        mean_value = account_history["Valor"].mean() if not account_history.empty else None
        max_row = account_history.loc[account_history["Valor"].idxmax()] if not account_history.empty else None

        cards = [
            {
                "title": f"Valor en {selected_year}",
                "value": format_compact_number(selected_value),
                "detail": short_label(selected_account, 56),
                "color": COLOR_POSITIVE,
            },
            {
                "title": f"Cambio vs {self.base_year}",
                "value": format_compact_number(delta_value),
                "detail": "Variacion absoluta de la cuenta seleccionada.",
                "color": COLOR_NEGATIVE if delta_value is not None and delta_value < 0 else COLOR_POSITIVE,
            },
            {
                "title": "Promedio 5 anos",
                "value": format_compact_number(mean_value),
                "detail": "Promedio simple del historial disponible.",
                "color": COLOR_ACCENT,
            },
            {
                "title": "Maximo historico",
                "value": format_compact_number(max_row["Valor"] if max_row is not None else None),
                "detail": f"Ano: {int(max_row['Anio'])}" if max_row is not None else "N/D",
                "color": COLOR_POSITIVE,
            },
        ]
        self._populate_cards(state["cards_frame"], cards)

        self._draw_figure(state["trend_section"], self._make_account_trend_figure(account_history, selected_account))
        self._draw_figure(state["ranking_section"], self._make_accounts_ranking_figure(filtered_wide, selected_year))

        detail_frame = filtered_wide[["Cuentas", *[str(year) for year in self.data.years]]].copy()
        detail_frame[f"Delta {self.latest_year}-{self.base_year}"] = detail_frame[str(self.latest_year)] - detail_frame[str(self.base_year)]
        self._draw_treeview(
            state["table_section"],
            detail_frame,
            compact_numeric_columns=set([str(year) for year in self.data.years] + [f"Delta {self.latest_year}-{self.base_year}"]),
            height=12,
        )