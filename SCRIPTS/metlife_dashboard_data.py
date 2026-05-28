"""Utilidades de carga y formateo para el dashboard standalone de MetLife."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries


CATEGORY_ORDER = [
    "Liquidez",
    "Solvencia y Apalancamiento",
    "Suficiencia de la Prima",
    "Reaseguro",
    "Rentabilidad",
]


@dataclass
class DashboardData:
    """Datos ya limpios y listos para mostrarse en la interfaz."""

    workbook_path: Path
    ratios_wide: pd.DataFrame
    ratios_long: pd.DataFrame
    accounts_wide: pd.DataFrame
    years: list[int]
    categories: list[str]


class ExcelDashboardLoader:
    """Lee el Excel del proyecto y arma la estructura que usa la app."""

    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"No se encontró el workbook: {self.workbook_path}")
        self.table_refs = self._resolve_named_tables()

    def load(self) -> DashboardData:
        # Paso 1: leer las tablas del Excel usando sus nombres, no rangos escritos a mano.
        ratios_wide = self._read_named_table("Razones_anuales")
        accounts_wide = self._read_named_table("Cuentas_anuales")

        # Paso 2: limpiar encabezados, vacíos y convertir columnas anuales a número.
        ratios_wide = self._clean_ratios_table(ratios_wide)
        accounts_wide = self._clean_accounts_table(accounts_wide)

        # Paso 3: detectar años y convertir razones al formato largo para las gráficas.
        years = self._detect_years(ratios_wide)
        year_columns = [str(year) for year in years]
        ratios_long = ratios_wide.melt(
            id_vars=["Categoria", "Razon"],
            value_vars=year_columns,
            var_name="Anio",
            value_name="Valor",
        )
        ratios_long["Anio"] = ratios_long["Anio"].astype(int)
        ratios_long["Valor"] = pd.to_numeric(ratios_long["Valor"], errors="coerce")
        ratios_long = ratios_long.dropna(subset=["Valor"]).reset_index(drop=True)

        # Paso 4: respetar el orden esperado de las categorías dentro del dashboard.
        workbook_categories = ratios_wide["Categoria"].dropna().astype(str).tolist()
        unique_categories = list(dict.fromkeys(workbook_categories))
        categories = [category for category in CATEGORY_ORDER if category in unique_categories]
        categories.extend([category for category in unique_categories if category not in categories])

        return DashboardData(
            workbook_path=self.workbook_path,
            ratios_wide=ratios_wide,
            ratios_long=ratios_long,
            accounts_wide=accounts_wide,
            years=years,
            categories=categories,
        )

    def _resolve_named_tables(self) -> dict[str, tuple[str, str]]:
        # Recorremos el libro para ubicar la hoja y el rango real de cada tabla nombrada.
        workbook = load_workbook(self.workbook_path, data_only=True)
        table_refs: dict[str, tuple[str, str]] = {}
        try:
            for sheet in workbook.worksheets:
                for table_name in sheet.tables.keys():
                    table_refs[table_name] = (sheet.title, sheet.tables[table_name].ref)
        finally:
            workbook.close()
        return table_refs

    def _read_named_table(self, table_name: str) -> pd.DataFrame:
        if table_name not in self.table_refs:
            available = ", ".join(sorted(self.table_refs.keys()))
            raise ValueError(f"No existe la tabla '{table_name}'. Tablas detectadas: {available}")

        # Convertimos la referencia de Excel a parámetros que pandas sí entiende.
        sheet_name, table_ref = self.table_refs[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table_ref)
        usecols = f"{get_column_letter(min_col)}:{get_column_letter(max_col)}"

        dataframe = pd.read_excel(
            self.workbook_path,
            sheet_name=sheet_name,
            header=min_row - 1,
            nrows=max_row - min_row,
            usecols=usecols,
        )
        dataframe = dataframe.dropna(how="all").dropna(axis=1, how="all")
        dataframe.columns = [self._normalize_column_name(column) for column in dataframe.columns]
        return dataframe

    @staticmethod
    def _normalize_column_name(column: object) -> str:
        # En la tabla de cuentas, pandas a veces trae 2021.1, 2022.1, etcétera.
        text = str(column).strip()
        match = re.fullmatch(r"(\d{4})(?:\.\d+)?", text)
        if match:
            return match.group(1)
        return text

    @staticmethod
    def _detect_years(dataframe: pd.DataFrame) -> list[int]:
        years = sorted(int(column) for column in dataframe.columns if re.fullmatch(r"\d{4}", str(column)))
        if not years:
            raise ValueError("No se detectaron columnas de años en la tabla de razones.")
        return years

    def _clean_ratios_table(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        # Aquí dejamos solo filas válidas y forzamos los años a formato numérico.
        required_columns = {"Categoria", "Razon"}
        if not required_columns.issubset(dataframe.columns):
            raise ValueError(f"La tabla de razones debe contener las columnas {sorted(required_columns)}")

        cleaned = dataframe.copy()
        cleaned["Categoria"] = cleaned["Categoria"].astype(str).str.strip()
        cleaned["Razon"] = cleaned["Razon"].astype(str).str.strip()
        cleaned = cleaned.loc[(cleaned["Categoria"] != "") & (cleaned["Razon"] != "")].reset_index(drop=True)

        for year in self._detect_years(cleaned):
            cleaned[str(year)] = pd.to_numeric(cleaned[str(year)], errors="coerce")

        return cleaned

    def _clean_accounts_table(self, dataframe: pd.DataFrame) -> pd.DataFrame:
        # Las cuentas ya no se grafican, pero se conservan para validar y resumir el archivo fuente.
        if "Cuentas" not in dataframe.columns:
            raise ValueError("La tabla de cuentas debe contener la columna 'Cuentas'.")

        cleaned = dataframe.copy()
        cleaned["Cuentas"] = cleaned["Cuentas"].astype(str).str.strip()
        cleaned = cleaned.loc[cleaned["Cuentas"] != ""].reset_index(drop=True)

        year_columns = [column for column in cleaned.columns if re.fullmatch(r"\d{4}", str(column))]
        if not year_columns:
            raise ValueError("La tabla de cuentas no contiene columnas anuales válidas.")

        for year_column in year_columns:
            cleaned[year_column] = pd.to_numeric(cleaned[year_column], errors="coerce")

        return cleaned


# Estas funciones mantienen el mismo formato visual en tarjetas, tablas y etiquetas.
def format_ratio(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    return f"{float(value):,.2f}"


def format_delta(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return "N/D"
    return f"{float(value):+,.2f}"


def short_label(text: str, limit: int = 42) -> str:
    cleaned = str(text).strip()
    if len(cleaned) <= limit:
        return cleaned
    return cleaned[: limit - 3].rstrip() + "..."
